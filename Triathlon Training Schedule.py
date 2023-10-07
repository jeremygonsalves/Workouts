import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
import datetime
import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry
from tkinter import ttk

times = [
    '6:00am', '6:30am', '7:00am', '7:30am', '8:00am', '8:30am', '9:00am', '9:30am', '10:00am', '10:30am', '11:00am',
    '11:30am', '12:00pm', '12:30pm', '1:00pm', '1:30pm', '2:00pm', '2:30pm', '3:00pm', '3:30pm', '4:00pm', '4:30pm',
    '5:00pm', '5:30pm', '6:00pm', '6:30pm', '7:00pm', '7:30pm', '8:00pm', '8:30pm', '9:00pm'
]

# Function to handle button click
def generate_calendar():
    workout_times = [[entry.get() for entry in dropdowns] for dropdowns in entries]

    # Get the start and end dates of the summer
    start_date = start_date_picker.get_date()
    end_date = end_date_picker.get_date()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Generate a sheet for each week of the summer
    current_date = start_date
    while current_date <= end_date:
        # Select the active sheet
        sheet = workbook.active

        # Set column headers
        sheet.cell(row=1, column=1).value = 'Time'
        for col, day in enumerate(days, start=2):
            sheet.cell(row=1, column=col).value = day

        # Set time blocks from 6am to 9pm with 30-minute intervals
        for row, time in enumerate(times, start=2):
            sheet.cell(row=row, column=1).value = time

        # Define fill colors for each activity type
        fill_colors = {
            'Long Run': 'FFFFFF00',  # Light Yellow
            'Workout': 'FFFFFF00',  # Light Yellow
            'Swim': 'FF00BFFF',  # Deep Sky Blue
            'Hard Bike': 'FFFFA500',  # Orange
            'Soccer': 'FFFFA500',  # Orange
            'Recovery Run': 'FFFFFF00',  # Light Yellow
            'Easy Run': 'FFFFFF00',  # Light Yellow
            'Brick Run': 'FFFFFF00'  # Light Yellow
        }

        # Populate the schedule with workout times and activity labels
        for i, day in enumerate(days):
            for j, activity in enumerate(activities[i]):
                col = days.index(day) + 2
                if i < len(workout_times) and j < len(workout_times[i]):
                    time = workout_times[i][j]
                    if time != "":
                        start_row = times.index(time) + 2
                        fill_color = fill_colors.get(activity, 'FF00FF00')  # Default to Light Green for unrecognized activities
                        cell = sheet.cell(row=start_row, column=col)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                        cell.font = Font(bold=True)  # Make the activity label bold
                        if j == 0:
                            cell.value = activity
                    else:
                        continue

        # Set column widths and row heights for better readability
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        for row_cells in sheet.rows:
            max_height = 0
            for cell in row_cells:
                try:
                    if len(str(cell.value)) > max_height:
                        max_height = len(cell.value)
                except:
                    pass
            adjusted_height = max_height + 6
            sheet.row_dimensions[row_cells[0].row].height = adjusted_height

        # Apply borders to all cells
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in sheet.iter_rows(min_row=1, max_row=len(times) + 1, min_col=1, max_col=len(days) + 1):
            for cell in row:
                cell.border = border

        # Set the sheet name as the specific week
        sheet.title = f"Week {current_date.strftime('%W')}"

        # Increment the current date by 7 days
        current_date += datetime.timedelta(days=7)

    # Save the workbook as an Excel file
    workbook.save("training_schedule.xlsx")
    messagebox.showinfo("Schedule Printed", "The training schedule has been printed to training_schedule.xlsx")
    window.destroy()

# Create the GUI window
window = tk.Tk()
window.title("Training Calendar")
window.geometry("500x500")

# Create date pickers for the start and end dates of the summer
start_date_label = tk.Label(window, text="Start Date:")
start_date_label.pack()
start_date_picker = DateEntry(window, date_pattern="dd/mm/yyyy")
start_date_picker.pack()

end_date_label = tk.Label(window, text="End Date:")
end_date_label.pack()
end_date_picker = DateEntry(window, date_pattern="dd/mm/yyyy")
end_date_picker.pack()

# Define the workout activities for each day
days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
activities = [
    ['Long Run'],
    ['Swim', 'Workout'],
    ['Tempo Run', 'Workout'],
    ['Hard Bike', 'Soccer'],
    ['Recovery Run', 'Workout'],
    ['Easy Run', 'Swim'],
    ['Long Bike', 'Brick Run']
]

# Create the dropdown menus for workout times
entries = []
for i, day in enumerate(days):
    label = tk.Label(window, text=day)
    label.pack()
    dropdowns = []
    for activity in activities[i]:
        time_label = tk.Label(window, text=activity + " Time:")
        time_label.pack()
        dropdown_var = tk.StringVar(window)
        dropdown = ttk.Combobox(window, textvariable=dropdown_var, values=times, state="readonly")
        dropdown.current(0)
        dropdown.pack()
        dropdowns.append(dropdown_var)
    entries.append(dropdowns)

# Create the button
button = tk.Button(window, text="Generate Schedule", command=generate_calendar)
button.pack()

# Start the GUI event loop
window.mainloop()
